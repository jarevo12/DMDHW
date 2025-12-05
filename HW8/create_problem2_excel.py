#!/usr/bin/env python3
"""Create Excel file for Problem 2 - Call Center Operator Scheduling."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_basic_model(wb):
    """Create Part (b) - Basic Model sheet."""
    ws = wb.create_sheet("Part B - Basic Model")

    # Title
    ws['A1'] = "Problem 2 Part (b): Call Center Operator Scheduling"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:N1')

    # Parameters section
    ws['A3'] = "PARAMETERS"
    ws['A3'].font = Font(size=12, bold=True, color="FFFFFF")
    ws['A3'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws.merge_cells('A3:N3')

    # Hour labels and requirements
    hours = ["7-8AM", "8-9AM", "9-10AM", "10-11AM", "11-12PM", "12-1PM",
             "1-2PM", "2-3PM", "3-4PM", "4-5PM", "5-6PM", "6-7PM"]
    requirements = [2, 3, 5, 8, 7, 5, 6, 7, 5, 5, 4, 4]
    rates = [25, 25, 18, 18, 18, 18, 18, 18, 18, 18, 25, 25]

    ws['A5'] = "Hour"
    ws['A5'].font = Font(bold=True)
    for i, hour in enumerate(hours, start=2):
        col = get_column_letter(i)
        ws[f'{col}5'] = hour
        ws[f'{col}5'].font = Font(bold=True, size=9)
        ws[f'{col}5'].alignment = Alignment(horizontal='center', text_rotation=90)

    ws['A6'] = "Min Required"
    ws['A6'].font = Font(bold=True)
    for i, req in enumerate(requirements, start=2):
        col = get_column_letter(i)
        ws[f'{col}6'] = req

    ws['A7'] = "Pay Rate ($/hr)"
    ws['A7'].font = Font(bold=True)
    for i, rate in enumerate(rates, start=2):
        col = get_column_letter(i)
        ws[f'{col}7'] = rate
        if rate == 25:
            ws[f'{col}7'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        else:
            ws[f'{col}7'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # Decision variables section - Working Hours
    ws['A9'] = "DECISION VARIABLES: WORKING HOURS (y_o,h)"
    ws['A9'].font = Font(size=12, bold=True, color="FFFFFF")
    ws['A9'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws.merge_cells('A9:O9')

    ws['A11'] = "Operator"
    ws['A11'].font = Font(bold=True)
    ws['N11'] = "Total"
    ws['N11'].font = Font(bold=True)
    ws['O11'] = "Required"
    ws['O11'].font = Font(bold=True)

    for i, hour in enumerate(hours, start=2):
        col = get_column_letter(i)
        ws[f'{col}11'] = i - 1  # Hour number
        ws[f'{col}11'].font = Font(bold=True, size=9)
        ws[f'{col}11'].alignment = Alignment(horizontal='center')

    # Operators A-J with decision variables
    operators = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
    for i, op in enumerate(operators, start=12):
        ws[f'A{i}'] = f"Op {op}"
        # Initialize all working hours to 0
        for h in range(12):
            col = get_column_letter(h + 2)
            ws[f'{col}{i}'] = 0
            ws[f'{col}{i}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        # Total hours formula
        ws[f'N{i}'] = f"=SUM(B{i}:M{i})"
        ws[f'O{i}'] = 7  # Required 7 hours

    # Coverage row (sum of operators per hour)
    ws['A22'] = "Coverage"
    ws['A22'].font = Font(bold=True)
    for h in range(12):
        col = get_column_letter(h + 2)
        ws[f'{col}22'] = f"=SUM({col}12:{col}21)"
        ws[f'{col}22'].font = Font(bold=True)
        ws[f'{col}22'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    ws['A23'] = "Required"
    ws['A23'].font = Font(bold=True)
    for i, req in enumerate(requirements, start=2):
        col = get_column_letter(i)
        ws[f'{col}23'] = req

    # Lunch variables section
    ws['A25'] = "LUNCH VARIABLES (L_o,h) - Hours 5,6,7 only (11AM-2PM)"
    ws['A25'].font = Font(size=12, bold=True, color="FFFFFF")
    ws['A25'].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws.merge_cells('A25:F25')

    ws['A27'] = "Operator"
    ws['B27'] = "11-12PM (h5)"
    ws['C27'] = "12-1PM (h6)"
    ws['D27'] = "1-2PM (h7)"
    ws['E27'] = "Total Lunch"
    ws['F27'] = "Required"
    for cell in ['A27', 'B27', 'C27', 'D27', 'E27', 'F27']:
        ws[cell].font = Font(bold=True)
        ws[cell].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    for i, op in enumerate(operators, start=28):
        ws[f'A{i}'] = f"Op {op}"
        ws[f'B{i}'] = 0
        ws[f'C{i}'] = 0
        ws[f'D{i}'] = 0
        ws[f'E{i}'] = f"=B{i}+C{i}+D{i}"
        ws[f'F{i}'] = 1  # Must take exactly 1 lunch
        for cell in [f'B{i}', f'C{i}', f'D{i}']:
            ws[cell].fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    # Cost calculation section
    ws['A40'] = "COST CALCULATION"
    ws['A40'].font = Font(size=12, bold=True, color="FFFFFF")
    ws['A40'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws.merge_cells('A40:E40')

    ws['A42'] = "Operator"
    ws['B42'] = "Work Hours"
    ws['C42'] = "Lunch Hours Paid"
    ws['D42'] = "Total Pay Hours"
    ws['E42'] = "Daily Cost"
    for cell in ['A42', 'B42', 'C42', 'D42', 'E42']:
        ws[cell].font = Font(bold=True)

    for i, op in enumerate(operators, start=43):
        op_idx = i - 43
        work_row = 12 + op_idx
        lunch_row = 28 + op_idx

        ws[f'A{i}'] = f"Op {op}"
        # Complex cost formula based on rate differentials
        # Off-peak: hours 1,2 (cols B,C) at $25
        # Peak: hours 3-10 (cols D-K) at $18
        # Off-peak: hours 11,12 (cols L,M) at $25
        # Lunch: always at $18
        ws[f'B{i}'] = f"=N{work_row}"  # Total work hours
        ws[f'C{i}'] = f"=E{lunch_row}"  # Lunch hours
        ws[f'D{i}'] = f"=B{i}+C{i}"  # Total paid hours

        # Cost formula: off-peak hours * 25 + peak hours * 18 + lunch * 18
        cost_formula = (
            f"=(B{work_row}+C{work_row})*25"  # Hours 1-2 (7-9AM) at $25
            f"+(D{work_row}+E{work_row}+F{work_row}+G{work_row}+H{work_row}+I{work_row}+J{work_row}+K{work_row})*18"  # Hours 3-10 (9AM-5PM) at $18
            f"+(L{work_row}+M{work_row})*25"  # Hours 11-12 (5-7PM) at $25
            f"+E{lunch_row}*18"  # Lunch at $18
        )
        ws[f'E{i}'] = cost_formula

    ws['A53'] = "TOTAL DAILY COST"
    ws['A53'].font = Font(bold=True, size=12)
    ws['E53'] = "=SUM(E43:E52)"
    ws['E53'].font = Font(bold=True, size=12)
    ws['E53'].fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")

    # Solver instructions
    ws['G40'] = "SOLVER SETUP INSTRUCTIONS"
    ws['G40'].font = Font(size=11, bold=True, color="FFFFFF")
    ws['G40'].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws.merge_cells('G40:K40')

    instructions = [
        "1. Objective: Minimize E53 (Total Cost)",
        "2. Variables:",
        "   - Working hours: B12:M21 (120 cells)",
        "   - Lunch: B28:D37 (30 cells)",
        "",
        "3. Constraints:",
        "   - Each operator works 7 hrs: N12:N21 = 7",
        "   - Each operator 1 lunch: E28:E37 = 1",
        "   - Min coverage: B22:M22 >= B23:M23",
        "   - Can't work during lunch:",
        "     For each operator row i (12-21):",
        "     F_i + B_(i+16) <= 1  (11-12PM)",
        "     G_i + C_(i+16) <= 1  (12-1PM)",
        "     H_i + D_(i+16) <= 1  (1-2PM)",
        "   - Span constraints:",
        "     B12+L12 <= 1, B12+M12 <= 1 (Op A)",
        "     Similar for all operators",
        "   - All variables binary {0,1}",
        "",
        "4. Method: Evolutionary/Branch-and-bound"
    ]

    for i, inst in enumerate(instructions, start=41):
        ws[f'G{i}'] = inst

    # Constraint checking section
    ws['G25'] = "CONSTRAINT CHECKS"
    ws['G25'].font = Font(size=11, bold=True, color="FFFFFF")
    ws['G25'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws.merge_cells('G25:J25')

    ws['G27'] = "Constraint"
    ws['H27'] = "Current"
    ws['I27'] = "Required"
    ws['J27'] = "Status"
    for cell in ['G27', 'H27', 'I27', 'J27']:
        ws[cell].font = Font(bold=True)

    for i in range(10):
        row = 28 + i
        op_row = 12 + i
        lunch_row = 28 + i
        ws[f'G{row}'] = f"Op {operators[i]} hours"
        ws[f'H{row}'] = f"=N{op_row}"
        ws[f'I{row}'] = 7
        ws[f'J{row}'] = f'=IF(H{row}=I{row},"OK","CHECK")'

    ws['G38'] = "Lunch checks"
    ws['G38'].font = Font(bold=True)
    ws['G38'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # Column widths
    ws.column_dimensions['A'].width = 12
    for col in range(2, 14):  # B-M
        ws.column_dimensions[get_column_letter(col)].width = 7
    ws.column_dimensions['N'].width = 8
    ws.column_dimensions['O'].width = 10
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 10

def create_excel_file():
    """Create complete Excel file for Problem 2."""

    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Create model
    create_basic_model(wb)

    # Save file
    output_file = "/workspaces/DMDHW/Problem2_Solution.xlsx"
    wb.save(output_file)
    print(f"Excel file created: {output_file}")

if __name__ == "__main__":
    create_excel_file()
