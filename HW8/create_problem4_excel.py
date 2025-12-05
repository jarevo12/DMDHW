#!/usr/bin/env python3
"""Create Excel file for Problem 4 - Concert Stage Setup Worker Assignment."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_basic_model(wb):
    """Create Part (b) - Basic Model sheet."""
    ws = wb.create_sheet("Part B - Basic Model")

    # Title
    ws['A1'] = "Problem 4 Part (b): Basic Worker Assignment Model"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:G1')

    # Parameters section
    ws['A3'] = "PARAMETERS"
    ws['A3'].font = Font(size=12, bold=True, color="FFFFFF")
    ws['A3'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws.merge_cells('A3:G3')

    # Worker data
    ws['A5'] = "Worker"
    ws['B5'] = "Hourly Rate"
    ws['C5'] = "Hours Available"
    ws['A5'].font = Font(bold=True)
    ws['B5'].font = Font(bold=True)
    ws['C5'].font = Font(bold=True)

    workers = [(1, 20, 20), (2, 18, 18), (3, 15, 15), (4, 12, 10)]
    for i, (worker, rate, hours) in enumerate(workers, start=6):
        ws[f'A{i}'] = f"Worker {worker}"
        ws[f'B{i}'] = rate
        ws[f'C{i}'] = hours

    # Task data
    ws['E5'] = "Task"
    ws['F5'] = "Description"
    ws['G5'] = "Hours Required"
    ws['E5'].font = Font(bold=True)
    ws['F5'].font = Font(bold=True)
    ws['G5'].font = Font(bold=True)

    ws['E6'] = "Task 1"
    ws['F6'] = "Structure building"
    ws['G6'] = 20
    ws['E7'] = "Task 2"
    ws['F7'] = "Electrical wiring"
    ws['G7'] = 15

    # Decision variables section
    ws['A11'] = "DECISION VARIABLES (Hours Worked)"
    ws['A11'].font = Font(size=12, bold=True, color="FFFFFF")
    ws['A11'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws.merge_cells('A11:E11')

    ws['A13'] = "Worker"
    ws['B13'] = "Structure (Task 1)"
    ws['C13'] = "Electrical (Task 2)"
    ws['D13'] = "Total Hours Used"
    ws['E13'] = "Capacity"
    for cell in ['A13', 'B13', 'C13', 'D13', 'E13']:
        ws[cell].font = Font(bold=True)
        ws[cell].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # Decision variable cells (initialize to 0)
    for i in range(4):
        row = 14 + i
        ws[f'A{row}'] = f"Worker {i+1}"
        ws[f'B{row}'] = 0  # x_{i,1}
        ws[f'C{row}'] = 0  # x_{i,2}
        ws[f'D{row}'] = f"=B{row}+C{row}"  # Total hours
        ws[f'E{row}'] = workers[i][2]  # Capacity

        # Format decision variables
        ws[f'B{row}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        ws[f'C{row}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # Task totals row
    ws['A18'] = "TOTAL HOURS"
    ws['A18'].font = Font(bold=True)
    ws['B18'] = "=SUM(B14:B17)"
    ws['C18'] = "=SUM(C14:C17)"
    ws['B18'].font = Font(bold=True)
    ws['C18'].font = Font(bold=True)
    ws['B18'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    ws['C18'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    ws['A19'] = "REQUIRED"
    ws['A19'].font = Font(bold=True)
    ws['B19'] = 20
    ws['C19'] = 15

    # Cost calculation section
    ws['A21'] = "COST CALCULATION"
    ws['A21'].font = Font(size=12, bold=True, color="FFFFFF")
    ws['A21'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws.merge_cells('A21:D21')

    ws['A23'] = "Worker"
    ws['B23'] = "Hours Used"
    ws['C23'] = "Rate"
    ws['D23'] = "Cost"
    for cell in ['A23', 'B23', 'C23', 'D23']:
        ws[cell].font = Font(bold=True)
        ws[cell].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    for i in range(4):
        row = 24 + i
        ws[f'A{row}'] = f"Worker {i+1} (${workers[i][1]}/hr)"
        ws[f'B{row}'] = f"=D{14+i}"
        ws[f'C{row}'] = workers[i][1]
        ws[f'D{row}'] = f"=B{row}*C{row}"

    ws['A28'] = "TOTAL COST"
    ws['A28'].font = Font(bold=True, size=12)
    ws['D28'] = "=SUM(D24:D27)"
    ws['D28'].font = Font(bold=True, size=12)
    ws['D28'].fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")

    # Constraints summary
    ws['F11'] = "CONSTRAINTS SUMMARY"
    ws['F11'].font = Font(size=12, bold=True, color="FFFFFF")
    ws['F11'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws.merge_cells('F11:H11')

    ws['F13'] = "Constraint"
    ws['G13'] = "Current"
    ws['H13'] = "Required"
    for cell in ['F13', 'G13', 'H13']:
        ws[cell].font = Font(bold=True)

    ws['F14'] = "Task 1 Hours"
    ws['G14'] = "=B18"
    ws['H14'] = "=B19"

    ws['F15'] = "Task 2 Hours"
    ws['G15'] = "=C18"
    ws['H15'] = "=C19"

    ws['F16'] = "Worker 1 Capacity"
    ws['G16'] = "=D14"
    ws['H16'] = "<=E14"

    ws['F17'] = "Worker 2 Capacity"
    ws['G17'] = "=D15"
    ws['H17'] = "<=E15"

    ws['F18'] = "Worker 3 Capacity"
    ws['G18'] = "=D16"
    ws['H18'] = "<=E16"

    ws['F19'] = "Worker 4 Capacity"
    ws['G19'] = "=D17"
    ws['H19'] = "<=E17"

    # Solver instructions
    ws['F21'] = "SOLVER SETUP"
    ws['F21'].font = Font(size=12, bold=True, color="FFFFFF")
    ws['F21'].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws.merge_cells('F21:H21')

    instructions = [
        "1. Objective: Minimize D28",
        "2. Variables: B14:C17",
        "3. Constraints:",
        "   - B18 = 20 (Task 1)",
        "   - C18 = 15 (Task 2)",
        "   - D14:D17 <= E14:E17",
        "   - B14:C17 >= 0",
        "   - B14:C17 = integer",
        "4. Method: Simplex LP"
    ]

    for i, inst in enumerate(instructions, start=22):
        ws[f'F{i}'] = inst

    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12

def create_extended_model(wb):
    """Create Part (d) - Extended Model with R1, R2, R3."""
    ws = wb.create_sheet("Part D - Extended Model")

    # Title
    ws['A1'] = "Problem 4 Part (d): Extended Model with Safety/Union Rules"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:H1')

    # Parameters (same as basic)
    ws['A3'] = "PARAMETERS"
    ws['A3'].font = Font(size=12, bold=True, color="FFFFFF")
    ws['A3'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws.merge_cells('A3:G3')

    ws['A5'] = "Worker"
    ws['B5'] = "Hourly Rate"
    ws['C5'] = "Hours Available"
    ws['A5'].font = Font(bold=True)
    ws['B5'].font = Font(bold=True)
    ws['C5'].font = Font(bold=True)

    workers = [(1, 20, 20), (2, 18, 18), (3, 15, 15), (4, 12, 10)]
    for i, (worker, rate, hours) in enumerate(workers, start=6):
        ws[f'A{i}'] = f"Worker {worker}"
        ws[f'B{i}'] = rate
        ws[f'C{i}'] = hours

    ws['E5'] = "Task"
    ws['F5'] = "Hours Required"
    ws['E5'].font = Font(bold=True)
    ws['F5'].font = Font(bold=True)

    ws['E6'] = "Task 1 (Structure)"
    ws['F6'] = 20
    ws['E7'] = "Task 2 (Electrical)"
    ws['F7'] = 15

    # Decision variables
    ws['A11'] = "DECISION VARIABLES"
    ws['A11'].font = Font(size=12, bold=True, color="FFFFFF")
    ws['A11'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws.merge_cells('A11:E11')

    ws['A13'] = "Worker"
    ws['B13'] = "Structure (Task 1)"
    ws['C13'] = "Electrical (Task 2)"
    ws['D13'] = "Total Hours"
    ws['E13'] = "Capacity"
    for cell in ['A13', 'B13', 'C13', 'D13', 'E13']:
        ws[cell].font = Font(bold=True)
        ws[cell].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    for i in range(4):
        row = 14 + i
        ws[f'A{row}'] = f"Worker {i+1}"
        ws[f'B{row}'] = 0
        ws[f'C{row}'] = 0
        ws[f'D{row}'] = f"=B{row}+C{row}"
        ws[f'E{row}'] = workers[i][2]
        ws[f'B{row}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        ws[f'C{row}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    ws['A18'] = "TOTAL HOURS"
    ws['A18'].font = Font(bold=True)
    ws['B18'] = "=SUM(B14:B17)"
    ws['C18'] = "=SUM(C14:C17)"
    ws['B18'].font = Font(bold=True)
    ws['C18'].font = Font(bold=True)
    ws['B18'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    ws['C18'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    # Binary variables section
    ws['A20'] = "BINARY INDICATOR VARIABLES"
    ws['A20'].font = Font(size=12, bold=True, color="FFFFFF")
    ws['A20'].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws.merge_cells('A20:E20')

    ws['A22'] = "Variable"
    ws['B22'] = "Value"
    ws['C22'] = "Description"
    for cell in ['A22', 'B22', 'C22']:
        ws[cell].font = Font(bold=True)

    binary_vars = [
        ("z1", "=IF(B14>0,1,0)", "Worker 1 works on structure (R1)"),
        ("z2", "=IF(B15>0,1,0)", "Worker 2 works on structure (R1)"),
        ("z3", "=IF(B16>0,1,0)", "Worker 3 works on structure (R1)"),
        ("z4", "=IF(B17>0,1,0)", "Worker 4 works on structure (R1)"),
        ("b3", 0, "Worker 3 task choice: 1=structure, 0=electrical (R2)"),
        ("u4", "=IF(C17>0,1,0)", "Worker 4 works on electrical (R3)")
    ]

    for i, (var, val, desc) in enumerate(binary_vars, start=23):
        ws[f'A{i}'] = var
        ws[f'B{i}'] = val
        ws[f'C{i}'] = desc
        ws[f'B{i}'].fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    # Cost calculation
    ws['A30'] = "COST CALCULATION"
    ws['A30'].font = Font(size=12, bold=True, color="FFFFFF")
    ws['A30'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws.merge_cells('A30:D30')

    ws['A32'] = "Worker"
    ws['B32'] = "Hours"
    ws['C32'] = "Rate"
    ws['D32'] = "Cost"
    for cell in ['A32', 'B32', 'C32', 'D32']:
        ws[cell].font = Font(bold=True)

    for i in range(4):
        row = 33 + i
        ws[f'A{row}'] = f"Worker {i+1}"
        ws[f'B{row}'] = f"=D{14+i}"
        ws[f'C{row}'] = workers[i][1]
        ws[f'D{row}'] = f"=B{row}*C{row}"

    ws['A37'] = "TOTAL COST"
    ws['A37'].font = Font(bold=True, size=12)
    ws['D37'] = "=SUM(D33:D36)"
    ws['D37'].font = Font(bold=True, size=12)
    ws['D37'].fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")

    # Additional constraints
    ws['F11'] = "ADDITIONAL CONSTRAINTS (R1, R2, R3)"
    ws['F11'].font = Font(size=12, bold=True, color="FFFFFF")
    ws['F11'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws.merge_cells('F11:I11')

    ws['F13'] = "Rule"
    ws['G13'] = "Constraint"
    ws['H13'] = "Current"
    ws['I13'] = "Required"
    for cell in ['F13', 'G13', 'H13', 'I13']:
        ws[cell].font = Font(bold=True)

    constraints = [
        ("R1", "At least 3 on structure", "=B23+B24+B25+B26", ">=3"),
        ("R1", "z1 links to x_1,1", "B14", "<=20*B23"),
        ("R1", "z2 links to x_2,1", "B15", "<=18*B24"),
        ("R1", "z3 links to x_3,1", "B16", "<=15*B25"),
        ("R1", "z4 links to x_4,1", "B17", "<=10*B26"),
        ("R2", "Worker 3: x_3,1 limit", "B16", "<=15*B27"),
        ("R2", "Worker 3: x_3,2 limit", "C16", "<=15*(1-B27)"),
        ("R3", "Worker 4: x_4,2 upper", "C17", "<=10*B28"),
        ("R3", "Worker 4: x_4,2 lower", "C17", ">=5*B28")
    ]

    for i, (rule, desc, current, required) in enumerate(constraints, start=14):
        ws[f'F{i}'] = rule
        ws[f'G{i}'] = desc
        ws[f'H{i}'] = current
        ws[f'I{i}'] = required

    # Solver instructions
    ws['F24'] = "SOLVER SETUP FOR EXTENDED MODEL"
    ws['F24'].font = Font(size=11, bold=True, color="FFFFFF")
    ws['F24'].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws.merge_cells('F24:I24')

    instructions = [
        "1. Objective: Minimize D37",
        "2. Variables: B14:C17, B27 (b3)",
        "3. Constraints (in addition to basic):",
        "   R1: B23+B24+B25+B26 >= 3",
        "   R1: B14 <= 20*B23, B15 <= 18*B24",
        "   R1: B16 <= 15*B25, B17 <= 10*B26",
        "   R2: B16 <= 15*B27",
        "   R2: C16 <= 15*(1-B27)",
        "   R3: C17 <= 10*B28",
        "   R3: C17 >= 5*B28",
        "4. B27 must be binary {0,1}",
        "5. Note: B23-B26, B28 auto-calculated"
    ]

    for i, inst in enumerate(instructions, start=25):
        ws[f'F{i}'] = inst

    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 22
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 15

def create_excel_file():
    """Create complete Excel file for Problem 4."""

    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Create both models
    create_basic_model(wb)
    create_extended_model(wb)

    # Save file
    output_file = "/workspaces/DMDHW/Problem4_Solution.xlsx"
    wb.save(output_file)
    print(f"Excel file created: {output_file}")

if __name__ == "__main__":
    create_excel_file()
